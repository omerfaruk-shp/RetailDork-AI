import asyncio
import time
import urllib.parse
import re
import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
from openai import OpenAI
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from ascii_magic import AsciiArt
import os 
def ascii_art_work():
    my_art = AsciiArt.from_image('meliodas.jpg')
    my_art.to_terminal()
    time.sleep(4)
    os.system('reset')
ascii_art_work()
LOCAL_API_URL = "http://localhost:11434/v1" 
client = OpenAI(base_url=LOCAL_API_URL, api_key="yerel-kullanim")

class UltimatePriceAndLinkScraper:
    def __init__(self, product_name):
        self.product_name = product_name
        self.results = []
        
        self.target_sites = [
            {"name": "Akakçe (Dip Fiyat)", "domain": "akakce.com"},
            {"name": "Cimri (Dip Fiyat)", "domain": "cimri.com"},
            {"name": "Epey", "domain": "epey.com"},
            {"name": "Trendyol", "domain": "trendyol.com"},
            {"name": "Hepsiburada", "domain": "hepsiburada.com"},
            {"name": "Amazon TR", "domain": "amazon.com.tr"},
            {"name": "N11", "domain": "n11.com"},
            {"name": "PttAVM", "domain": "pttavm.com"},
            {"name": "Pazarama", "domain": "pazarama.com"},
            {"name": "Çiçeksepeti", "domain": "ciceksepeti.com"},
            {"name": "Teknosa", "domain": "teknosa.com"},
            {"name": "Vatan Bilgisayar", "domain": "vatanbilgisayar.com"}
        ]
        
        # İkinci el ve aksesuar engelleme filtresi
        self.noise_keywords = [
            "kılıf", "kaplama", "çantası", "çanta", "koruyucu", "sticker", "etiket", 
            "analog koruyucu", "stand", "stanti", "yedek parça", "tamir", "kablo", 
            "silikon", "skin", "tutucu", "grip", "askı", "cam",
            "letgo", "ikinci el", "kullanılmış", "deforme", 
            "az kullanılmış", "yenilenmiş", "teşhir", "outlet"
        ]

    def _is_relevant_product(self, title):
        """Ürünün aranan kelimeyle alakalı ve SIFIR olduğunu doğrular."""
        title_lower = title.lower()
        query_lower = self.product_name.lower()
        
        query_words = [w for w in query_lower.split() if len(w) > 2]
        if query_words:
            matched = any(word in title_lower for word in query_words)
            if not matched:
                return False

        for noise in self.noise_keywords:
            if noise in title_lower and noise not in query_lower:
                return False
                
        return True

    def _extract_clean_url(self, element):
        """Google'ın gizlediği veya yönlendirdiği gerçek ürün URL'sini çıkarır."""
        candidates = []
        
        if element.name == 'a' and element.has_attr('href'):
            candidates.append(element['href'])
        
        for a_tag in element.select('a[href]'):
            candidates.append(a_tag['href'])

        for href in candidates:
            if not href:
                continue
            
            if any(ignored in href for ignored in ['support.google.com', 'google.com/about', 'google.com/policies', 'accounts.google.com', 'search?q=']):
                continue

            if '/url?' in href or 'adurl=' in href:
                parsed = urllib.parse.urlparse(href)
                qs = urllib.parse.parse_qs(parsed.query)
                if 'q' in qs and qs['q'][0].startswith('http'):
                    return qs['q'][0]
                elif 'adurl' in qs and qs['adurl'][0].startswith('http'):
                    return qs['adurl'][0]

            if href.startswith('http') and 'google.com' not in href:
                return href
            elif href.startswith('/shopping/product/'):
                return f"https://www.google.com{href}"

        return ""

    def _extract_price(self, text):
        """Metin içerisinden TL/₺ fiyat formatlarını ayıklar."""
        match = re.search(r'(\d{1,3}(?:\.\d{3})*(?:,\d{1,2})?)\s*(?:TL|₺)', text)
        if match:
            return f"{match.group(1)} TL"
        
        match_alt = re.search(r'(?:TL|₺)\s*(\d{1,3}(?:\.\d{3})*(?:,\d{1,2})?)', text)
        if match_alt:
            return f"{match_alt.group(1)} TL"
            
        return "Fiyat Belirsiz"

    async def scrape_google_web_dorks(self, page, site_info):
        """Google Web Arama Dorkları (Akakçe, Cimri, Trendyol vb. için en ucuz sonuçlar)."""
        site_name = site_info["name"]
        domain = site_info["domain"]
        dork_query = f'site:{domain} "{self.product_name}"'
        
        print(f"🔍 [DORK TARAMA] {site_name} -> '{dork_query}'")
        encoded_query = urllib.parse.quote_plus(dork_query)
        target_url = f"https://www.google.com/search?q={encoded_query}&num=10"

        try:
            await page.goto(target_url, wait_until="domcontentloaded", timeout=15000)
            await asyncio.sleep(1)

            html_content = await page.content()
            soup = BeautifulSoup(html_content, 'html.parser')
            
            search_results = soup.select('#rso .g, .MjjYud, .tF2C3e')

            extracted = 0
            for card in search_results:
                title_tag = card.select_one('h3')
                if not title_tag:
                    continue
                title = title_tag.get_text(strip=True)

                if not self._is_relevant_product(title):
                    continue

                product_link = self._extract_clean_url(card)
                if not product_link:
                    continue

                card_text = card.get_text()
                price = self._extract_price(card_text)

                key = (title, site_name)
                if key not in [(r["Ürün Başlığı"], r["Platform / Satıcı"]) for r in self.results]:
                    self.results.append({
                        "Platform / Satıcı": site_name,
                        "Ürün Başlığı": title,
                        "Fiyat": price,
                        "Puan": "Web İlanı",
                        "Ürün Linki": product_link
                    })
                    extracted += 1

            print(f"   ↳ {extracted} geçerli link ve fiyat bulundu.")

        except Exception as e:
            print(f"   ⚠️ {site_name} taranırken hata: {e}")

    async def scrape_google_shopping_feed(self, page, max_pages=3):
        """Google Shopping Genel Arama (udm=28)."""
        print(f"\n🛒 [GOOGLE SHOPPING FEED] Taranıyor...")
        query = f"{self.product_name} en ucuz sıfır"
        encoded_query = urllib.parse.quote_plus(query)

        for page_num in range(max_pages):
            start_offset = page_num * 10
            target_url = f"https://www.google.com/search?q={encoded_query}&udm=28&start={start_offset}"
            
            try:
                await page.goto(target_url, wait_until="domcontentloaded", timeout=15000)
                
                for _ in range(3):
                    await page.mouse.wheel(0, 800)
                    await asyncio.sleep(0.4)

                html_content = await page.content()
                soup = BeautifulSoup(html_content, 'html.parser')
                cards = soup.select('.sh-dgr__grid-result, div[data-docid], .Ez5pwe, .sh-np__click-target')

                extracted = 0
                for card in cards:
                    title_tag = card.select_one('h3, .gkQHve, .SsM98d, .RmEs5b')
                    if not title_tag:
                        continue
                    title = title_tag.get_text(strip=True)

                    if not self._is_relevant_product(title):
                        continue

                    price = self._extract_price(card.get_text())
                    
                    seller = "Google Shopping"
                    seller_tag = card.select_one('span.WJMUdc, .rw5ecc, .aULB2b, .a8f25')
                    if seller_tag:
                        seller = seller_tag.get_text(strip=True)

                    product_link = self._extract_clean_url(card)
                    if not product_link:
                        continue

                    key = (title, price)
                    if key not in [(r["Ürün Başlığı"], r["Fiyat"]) for r in self.results]:
                        self.results.append({
                            "Platform / Satıcı": seller,
                            "Ürün Başlığı": title,
                            "Fiyat": price,
                            "Puan": "Alışveriş Sekmesi",
                            "Ürün Linki": product_link
                        })
                        extracted += 1

                print(f"   ↳ Sayfa {page_num + 1}: {extracted} ürün eklendi.")
                await asyncio.sleep(0.5)

            except Exception as e:
                print(f"   ⚠️ Shopping Sayfa {page_num + 1} hatası: {e}")

    async def run_scraper(self):
        print("="*60)
        print(f"🚀 ULTRA UCUZ FİYAT VE LİNK ARAMA MOTORU BAŞLATILDI")
        print(f"🎯 Hedef Ürün: '{self.product_name}'")
        print("="*60)

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                viewport={'width': 1280, 'height': 900},
                locale='tr-TR'
            )
            page = await context.new_page()
            await page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

            init_url = f"https://www.google.com/search?q={urllib.parse.quote_plus(self.product_name)}"
            await page.goto(init_url, wait_until="domcontentloaded", timeout=20000)

            print("\n⚠️ BOT DOĞRULAMA KONTROLÜ:")
            print("1. Tarayıcıda bot kontrolü (reCAPTCHA) çıkarsa tamamlayın.")
            print("2. Sayfa açıldığında terminale gelip ENTER'a basın!\n")
            input("Taramayı başlatmak için ENTER tuşuna basın...")

            print("\n📌 [1. AŞAMA] Fiyat Karşılaştırma & Mağaza Dork Taraması")
            for site in self.target_sites:
                await self.scrape_google_web_dorks(page, site)

            print("\n📌 [2. AŞAMA] Google Shopping Feed Taraması")
            await self.scrape_google_shopping_feed(page, max_pages=3)

            await browser.close()

    def _parse_price_to_number(self, price_str):
        if not price_str or "Belirsiz" in price_str:
            return float('inf')
        clean_str = re.sub(r'[^\d,]', '', str(price_str))
        clean_str = clean_str.replace(',', '.')
        try:
            return float(clean_str)
        except ValueError:
            return float('inf')

    def export_to_excel(self, df):
        safe_name = re.sub(r'[\\/*?:"<>|]', "", self.product_name).replace(" ", "_").lower()
        file_name = f"{safe_name}_ucuz_fiyat_listesi.xlsx"

        # Fiyata göre artan sıralama (En ucuz ürün en üstte)
        df['Fiyat_Sayısal'] = df['Fiyat'].apply(self._parse_price_to_number)
        df = df.sort_values(by='Fiyat_Sayısal', ascending=True).reset_index(drop=True)
        df = df.drop(columns=['Fiyat_Sayısal'])

        try:
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Fiyat Listesi')
                worksheet = writer.sheets['Fiyat Listesi']

                for col in worksheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        val_str = str(cell.value or '')
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    calculated_width = max(max_len + 4, 12)
                    worksheet.column_dimensions[col_letter].width = min(calculated_width, 50)

                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                link_col_idx = df.columns.get_loc("Ürün Linki") + 1
                link_font = Font(name="Arial", size=10, color="0000FF", underline="single")

                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=link_col_idx)
                    url = cell.value
                    if url and str(url).startswith("http"):
                        cell.value = "Tıkla & Git 🔗"
                        cell.hyperlink = url
                        cell.font = link_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1):
                    worksheet.row_dimensions[row[0].row].height = 22

            print("\n" + "="*60)
            print(f"📊 EXCEL DOSYASI BAŞARIYLA HAZIRLANDI!")
            print(f"📦 Bulunan Ürün Sayısı: {len(df)}")
            print(f"📁 Dosya Yolu: /Users/faruk/{file_name}")
            print("="*60)

        except Exception as e:
            print(f"\n⚠️ Excel dosyası yazılırken hata: {e}")

    def analyze_with_local_ai(self, df):
        print("\n🧠 Yerel Gemma Modeli Dip Fiyat Analizi Yapıyor...")
        
        df_for_ai = df.drop(columns=['Ürün Linki'], errors='ignore')
        sample_df = df_for_ai.head(30)
        data_text = sample_df.to_string(index=False)
        
        prompt = f"""
        Aşağıda '{self.product_name}' ürünü için toplanan en uygun fiyatlı ilanlar yer alıyor:
        
        {data_text}
        
        Lütfen verileri analiz et:
        1. En dip fiyatı sunan satıcı/platform hangisi?
        2. Akakçe/Cimri ve mağazalar arasındaki fiyat farkı ne kadar?
        3. Fiyat/Performans açısından en mantıklı seçenek hangisi?
        
        Kısa ve net Türkçe bir rapor sun.
        """

        try:
            response = client.chat.completions.create(
                model="gemma4:12b", 
                messages=[{"role": "user", "content": prompt}],
                temperature=0.7
            )
            content = response.choices[0].message.content
            if content:
                print("\n🤖   GEMMA  ANALİZ RAPORU:")
                print("-" * 50)
                print(content)
                print("-" * 50)
        except Exception as e:
            print(f"Yerel AI rapor üretemedi, ancak Excel dosyan hazır. Hata: {e}")

    async def run(self):
        start_time = time.time()
        await self.run_scraper()
        
        df = pd.DataFrame(self.results)
        if not df.empty:
            df = df.drop_duplicates(subset=['Ürün Başlığı', 'Platform / Satıcı']).reset_index(drop=True)
            self.export_to_excel(df)
            self.analyze_with_local_ai(df)
        else:
            print("\nHiçbir sonuç bulunamadı.")
        
        elapsed = time.time() - start_time
        print(f"\n⏱️ İşlem süresi: {elapsed:.2f} saniye.")

if __name__ == "__main__":
    aranacak_urun = input("Aratılacak Ürün / Model: ")
    analyzer = UltimatePriceAndLinkScraper(aranacak_urun)
    asyncio.run(analyzer.run())
